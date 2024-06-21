#se importan las librerias que se utilizaran para el proyecto
import requests
from bs4 import BeautifulSoup
import openpyxl

#se hace la peticion al sitio web y se convierte en un texto con el que s epodra trabajar
website = "https://mascaradelatex.com/collections/hombre"
result = requests.get(website)
soup = BeautifulSoup(result.content, 'html.parser')

#arreglos donde se guardara la información extraida
nombre = []
price = []

#se encuentran todos los objetos que coincidan con lo establecido
shirts_names = soup.find_all('h3', class_ = 'c-card-product__title')
shirts_prices = soup.find_all('p', class_ = 'c-card-product__price')

#se guarda el contenido de lo que se obtuvo en nombre
for h3 in shirts_names:
  nombre.append(h3.text.strip())

#se guarda el contenido de lo que se obtuvo en price
for p in shirts_prices:
  price.append(p.text.strip())

#se abre un libro para poder crear un excell
workbook = openpyxl.Workbook()

#se crea la hoja en el libro de excell
hoja = workbook.active

#se guarda la información en 2 columnas de excel
for i in range(len(nombre)):
  hoja.cell(row = i + 1, column = 1, value = nombre[i])
  hoja.cell(row = i + 1, column = 2, value = price[i])

workbook.save("WebScrapping.xlsx")